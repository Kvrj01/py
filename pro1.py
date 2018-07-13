from urllib.request import urlopen
from bs4 import BeautifulSoup
import xlsxwriter
from openpyxl import load_workbook
from xlsxwriter.utility import xl_rowcol_to_cell
from xlsxwriter.utility import xl_cell_to_rowcol
from openpyxl.styles import PatternFill
import openpyxl.styles
import re
d = {}
f1=open(r'B:\hello.txt','r+')
line=f1.readlines()
for url in line:
    # request to open the url in order to get the html contents
    html = urlopen(url)
    # Scrap the contents using beautifulSoup
    r = BeautifulSoup(html.read(), 'html5lib');
    lb = len(r('body'))
    lt = len(r('title'))

    j = 0
    list1 = []
    list3 = []
    for i in range(0, lb):
        a = r.find_all('body')[j].get_text()
        list1.append(a.lower())
        j = j + 1

    # print(list1)
    for i in list1:
        list2 = i.split()
    # print(list2)

    a = r.find('title').get_text()
    list3.append(a.lower())
    l_fin = list2 + list3
    #print(l_fin)

    len_l_fin = len(l_fin)
    l_final =[]

    for i in l_fin:
        s=re.findall(r'[a-zA-z]+[a-zA-Z]',i)
        l_final.extend(s)

    for x in l_final:
        cnt = l_final.count(x)
        d2 = {x: cnt}
        d.update(d2)
#print(d)
# copying into excel

wbk= xlsxwriter.Workbook(r'B:\w2.xlsx')
wsk=wbk.add_worksheet()
t=d.items()
l=list(t)
#print(l)

# arranging the contents in descending order
def fun1(x):
    return x[1]
l.sort(key= fun1,reverse=True)

r=0
c=0
for i,j in l:
    wsk.write(r,c, i)
    wsk.write(r,c+1, j)
    r=r+1

chart = wbk.add_chart({'type': 'bar'})

# Add a series to the chart.
chart.add_series({
                   'categories': '=Sheet1!$A$1:$A$6',
                   'values': '=Sheet1!$B$1:$B$6'})

# Insert the chart into the worksheet.
wsk.insert_chart('C1', chart)

wbk.close()

str1=input('Enter a word to be serached: ')
f=0
wb = load_workbook(r'B:\w2.xlsx')
sheet = wb['Sheet1']


for i in range(1, sheet.max_row + 1):
    for j in range(sheet.max_column):
        if(str1.lower()==sheet[i][j].value):
            print(sheet[i][j].value ,':', sheet[i][j+1].value)
            celli = xl_rowcol_to_cell((i-1),j)
            print(celli)
            cellj = xl_rowcol_to_cell((i - 1), (j+1))
            print(cellj)
            (row1, col1) = xl_cell_to_rowcol(celli)
            #print(row1,col1)
            sheet['A1'].fill = openpyxl.styles.PatternFill('solid', openpyxl.styles.colors.GREEN)
            f=1
            break
if (f==0):
    print('Word not found')
wb.save('w2.xlsx')
wb.close()
f1.close()
